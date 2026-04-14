import openpyxl as xl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from typing import List


line_thin = Side(style="thin")

line_left = Border(left=line_thin)
line_right = Border(right=line_thin)
line_top = Border(top=line_thin)
line_bottom = Border(bottom=line_thin)
line_around = Border(left=line_thin, right=line_thin, top=line_thin, bottom=line_thin)
alignment_center = Alignment(horizontal='center')
alignment_left = Alignment(horizontal='left')

col_time = 1
col_op_nr = 2
col_title1 = 3
col_title2 =4
col_method = 4
col_content = 5
col_record = 6
col_operator = 7
col_witness = 8


class Flowsheet:

    def __init__(self):
        self.wb = xl.Workbook()
        self.ws = self.wb.active
        self.current_line = 1


    
    def header_organizer(self, op_nr: int, title: str):
        self.ws.merge_cells(start_row=self.current_line, start_column=col_title1, end_row=self.current_line, end_column=col_title2)
        self.ws.cell(row=self.current_line, column=col_op_nr, value=op_nr)
        self.ws.cell(row=self.current_line, column=col_title1, value=title)
        self.ws.cell(row=self.current_line, column=col_op_nr).border = line_around
        self.ws.cell(row=self.current_line, column=col_op_nr).alignment = alignment_center
        self.ws.cell(row=self.current_line, column=col_title1).border = Border(left=line_thin, top=line_thin, bottom=line_thin)
        self.ws.cell(row=self.current_line, column=col_title2).border = Border(top=line_thin, bottom=line_thin, right=line_thin)
        self.ws.cell(row=self.current_line, column=col_title1).alignment = alignment_center
        self.current_line += 1

    def body_organizer(self, list_col_time: List[str], list_col_method: List[str], list_col_content: List[str], list_col_record: List[str], list_col_operator: List[str], list_col_witness: List[str]):
        len_list_time = len(list_col_time)
        for row_rel in range(len_list_time):
            self.ws.cell(row=self.current_line+row_rel, column=col_time).value = list_col_time[row_rel]

        len_list_method = len(list_col_method)
        for row_rel in range(len_list_method):
            self.ws.cell(row=self.current_line+row_rel, column=col_method).value = list_col_method[row_rel]

        len_list_content = len(list_col_content)
        for row_rel in range(len_list_content):
            self.ws.cell(row=self.current_line+row_rel, column=col_content).value = list_col_content[row_rel]

        len_list_record = len(list_col_record)
        for row_rel in range(len_list_record):
            self.ws.cell(row=self.current_line+row_rel, column=col_record).value = list_col_record[row_rel]

        len_list_operator = len(list_col_operator)
        for row_rel in range(len_list_operator):
            self.ws.cell(row=self.current_line+row_rel, column=col_operator).value = list_col_operator[row_rel]

        len_list_witness = len(list_col_witness)
        for row_rel in range(len_list_witness):
            self.ws.cell(row=self.current_line+row_rel, column=col_witness).value = list_col_witness[row_rel]

        list_col_time.clear()
        list_col_method.clear()
        list_col_content.clear()
        list_col_record.clear()
        list_col_operator.clear()
        list_col_witness.clear()

        max_length = max(len_list_time,
                         len_list_method,
                         len_list_content,
                         len_list_record,
                         len_list_operator,
                         len_list_witness)
        self.current_line += max_length
    
    def linefeed(self):
        self.current_line += 1




    def test_save(self):
        self.wb.save("test_output.xlsx")




    