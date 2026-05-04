import openpyxl as xl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from typing import List
import re
import flow_draw.definitions as defs


line_thin = defs.xl_line_thin

border_left = defs.xl_border_left
border_right = defs.xl_border_right
border_top = defs.xl_border_top
border_bottom = defs.xl_border_bottom
border_around = defs.xl_border_around
alignment_center = defs.xl_alignment_center
alignment_left = defs.xl_alignment_left

col_time = 1
col_op_nr = 2
col_title_left_half = 3
col_title_right_half =4
col_method = 4
col_content = 5
col_record = 6
col_operator = 7
col_witness = 8


class Flowsheet:

    def __init__(self):
        self.wb = xl.Workbook()
        self.ws = self.wb.active
        self.current_line = 1

    def put_body_comments(self, comments :str):
        cmt_brkdwn = re.split('[\n;]', comments)
        self.body_organizer(list_col_time=[],
                            list_col_method=[],
                            list_col_content=cmt_brkdwn,
                            list_col_record=[],
                            list_col_operator=[],
                            list_col_witness=[])
        #self.linefeed()

    def header_organizer(self, op_nr: int, title: str):
        self.ws.merge_cells(start_row=self.current_line, start_column=col_title_left_half, end_row=self.current_line, end_column=col_title_right_half)
        self.ws.cell(row=self.current_line, column=col_op_nr, value=op_nr)
        self.ws.cell(row=self.current_line, column=col_title_left_half, value=title)
        self.ws.cell(row=self.current_line, column=col_op_nr).border = border_around
        self.ws.cell(row=self.current_line, column=col_op_nr).alignment = alignment_center
        self.ws.cell(row=self.current_line, column=col_title_left_half).border = Border(left=line_thin, top=line_thin, bottom=line_thin)
        self.ws.cell(row=self.current_line, column=col_title_right_half).border = Border(top=line_thin, bottom=line_thin, right=line_thin)
        self.ws.cell(row=self.current_line, column=col_title_left_half).alignment = alignment_center
        self.current_line += 1

    def put_line(self, time: str ='', method: str='', content: str='', record: str='', operator: str='', witness: str=''):
        self.ws.cell(row=self.current_line, column=col_time).value = time
        self.ws.cell(row=self.current_line, column=col_method).value = method
        self.ws.cell(row=self.current_line, column=col_content).value = content
        self.ws.cell(row=self.current_line, column=col_record).value = record
        self.ws.cell(row=self.current_line, column=col_operator).value = operator
        self.ws.cell(row=self.current_line, column=col_witness).value = witness
        self.ws.cell(row=self.current_line, column=col_title_left_half).border = border_left
        self.current_line += 1

    def body_organizer(self, list_col_time: List[str], list_col_method: List[str], list_col_content: List[str], list_col_record: List[str], list_col_operator: List[str], list_col_witness: List[str]):
        len_list_time = len(list_col_time)
        for row_rel in range(len_list_time):
            self.ws.cell(row=self.current_line+row_rel, column=col_time).value = list_col_time[row_rel]

        len_list_method = len(list_col_method)
        for row_rel in range(len_list_method):
            self.ws.cell(row=self.current_line+row_rel, column=col_method).value = list_col_method[row_rel]

        len_list_content = len(list_col_content)
        for row_rel in range(len_list_content):
            self.ws.cell(row=self.current_line+row_rel, column=col_content).value = list_col_content[row_rel]

        len_list_record = len(list_col_record)
        for row_rel in range(len_list_record):
            self.ws.cell(row=self.current_line+row_rel, column=col_record).value = list_col_record[row_rel]

        len_list_operator = len(list_col_operator)
        for row_rel in range(len_list_operator):
            self.ws.cell(row=self.current_line+row_rel, column=col_operator).value = list_col_operator[row_rel]

        len_list_witness = len(list_col_witness)
        for row_rel in range(len_list_witness):
            self.ws.cell(row=self.current_line+row_rel, column=col_witness).value = list_col_witness[row_rel]

        list_col_time.clear()
        list_col_method.clear()
        list_col_content.clear()
        list_col_record.clear()
        list_col_operator.clear()
        list_col_witness.clear()

        max_length = max(len_list_time,
                         len_list_method,
                         len_list_content,
                         len_list_record,
                         len_list_operator,
                         len_list_witness)
        
        for row_rel in range(max_length):
            self.ws.cell(row=self.current_line+row_rel, column=col_title_left_half).border = border_left

        self.current_line += max_length
    
    def linefeed(self):
        self.ws.cell(row=self.current_line, column=col_title_left_half).border = border_left
        self.current_line += 1




    def save(self, filename: str):
        self.wb.save(filename=filename)




    