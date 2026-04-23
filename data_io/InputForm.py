import os
import math
import openpyxl as xl
import pandas as pd
import flow_draw.definitions as defs
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation
from typing import List, Dict

base_file_name_input = "input_"
sheet_summary_input = "summary"
sheet_detail_input = "detail"

header_summary_sequence = 'Sequence'
header_summary_uo = 'Unit Operation'
header_summary_num_subitems = 'Number of Subitems'
header_summary_edit_comment = 'Edit Comment'

summary_col_seq = 1
summary_col_uo = 2
summary_col_num_subitems = 3
summary_col_editcomment = 4

header_detail_seq = header_summary_sequence
header_detail_operation = header_summary_uo
header_detail_edit_comment = header_summary_edit_comment
header_detail_precomment = 'Pre-comment'
header_detail_postcomment = 'Post-comment'

common_header_detail = [
    header_detail_seq,
    header_detail_operation,
    header_detail_edit_comment,
    header_detail_precomment,
    header_detail_postcomment 
]

no_comment_instr = '(No comment here)'

class InputForm:
    """
    This class manages IO for process information. Each instance of this class corresponds one-on-one with a process.
    This class has methods to create an input form for summary and details each, load data from the forms, and save the forms as worksheets in an Excel workbook.
    
    Attributes
    -------------
    process_name : str
        The name of the process.
    count_unit_op : int
        The number of unit operations that constitute the process. This must be given from outside the class at first in order to generate the form with proper number of unit operations.
    file_path : str
        The path to the data input form (Excel). The name is automatically generated from the process name. 
    summary_ws : openpyxl.worksheet.worksheet.Worksheet
        Worksheet for summary input.
    detail_ws : openpyxl.worksheet.worksheet.Worksheet
        Worksheet for detail input.
    df_summary : Pandas.DataFrame
        A narrow table for summary input. Holds only sequence numbers, unit operations, and edit comments. 
    _current_line_summary : int
        Current reading/writing line for the summary worksheet. For internal use only.
    _current_line_detail : int
        Current reading/writing line for the detail worksheet. Completely internal.


    """
    def __init__(self, process_name: str, num_unit_op: int):
        self.process_name:str = process_name
        self.count_unit_op: int = num_unit_op
        self.file_path = base_file_name_input+process_name+'.xlsx'
        self.summary_ws: Worksheet = None
        self.detail_ws: Worksheet = None
        self.__manage_io()
        self.df_summary: pd.DataFrame = None
        self._current_line_summary:int = 1
        self._current_line_detail:int = 1
    
    def __manage_io(self):
        """
        Creates an Open PyXL workbook for data input, and creates brank work sheets for summary and detail input, if the data input file with the intended name doesn't eixists.
        Otherwise, the files is loaded to re-generate an Open PyXL workbook. Worksheets are created as necessary.
        Please note this only takes care of worksheets, and doesn't handle DataFrame. It's up to other methods such as load_process_summary() or load_process_details().
        """
        if not os.path.isfile(self.file_path):
            self.wb = xl.Workbook()
            self.wb.remove(worksheet=self.wb['Sheet'])
            self.summary_ws:Worksheet = self.wb.create_sheet(title=sheet_summary_input)
            self.detail_ws: Worksheet = self.wb.create_sheet(title=sheet_detail_input)
        else:
            self.wb=xl.load_workbook(self.file_path)
            sheet_names = self.wb.sheetnames
            if sheet_names.count(sheet_summary_input) == 0:
                self.summary_ws:Worksheet = self.wb.create_sheet(title=sheet_summary_input)
            else:
                self.summary_ws: Worksheet = self.wb[sheet_summary_input]
            if sheet_names.count(sheet_detail_input) == 0:
                self.detail_ws: Worksheet = self.wb.create_sheet(title=sheet_detail_input)
            else:
                self.detail_ws: Worksheet = self.wb[sheet_detail_input]



    def save_form(self):
        """
        Saves the workbook which holds the worksheets for both summary and details. The output is an Excel file on the storage. No arguments.
        
        Parameters
        -----------
        None

        Returns
        ----------
        None
        """
        self.wb.save(filename=self.file_path)

    def get_paht_to_forms(self)->str:
        return self.file_path

    def put_summary_input_form(self, list_unit_ops: List[str]):
        """
        Creates the summary input form based on the number of the unit operations in the process.\n
        There will be four coloumns in the newly created form:\n
            -Sequnece Number
            -Unit Operation
            -Nuber of Sub-items
            -Edit Comment
        A pull-down menu will be set in each cell in the Unit Operation column.
        The method only edits the summary_ws in the class, and does not save it as an file on the storage. Another method must be called to do so.

        Parameters
        ---------------
        list_unit_ops: List[str]
            A list of strings which holds the name tags for unit operations, e.g., line_clearance, N2_replacement_ charging, etc. The name tags must be recognizable by the unit_operations.UnitOperation class.
        
        Returns
        ---------------
        None

        Notes
        ---------------
        Requires self.summary_ws: openpyxl.worksheet.worksheet.Worksheet. The worksheet must be available before this method is called. 

        """
        options_dv: str = '"'
        for item in list_unit_ops:
            options_dv += (item+',')
        options_dv = options_dv[:-1]
        options_dv = options_dv+'"'
        dv_unitops = DataValidation(
            type='list',
            formula1=options_dv,
            allow_blank=False
        )
        self.summary_ws.add_data_validation(dv_unitops)

        self._current_line_summary = 1
        self.summary_ws.cell(row = self._current_line_summary, column = summary_col_seq, value=header_summary_sequence)
        self.summary_ws.cell(row = self._current_line_summary, column = summary_col_seq).border = defs.xl_border_around
        self.summary_ws.cell(row = self._current_line_summary, column = summary_col_uo, value=header_summary_uo)
        self.summary_ws.cell(row = self._current_line_summary, column = summary_col_uo).border = defs.xl_border_around
        self.summary_ws.cell(row = self._current_line_summary, column = summary_col_num_subitems, value=header_summary_num_subitems)
        self.summary_ws.cell(row = self._current_line_summary, column = summary_col_num_subitems).border = defs.xl_border_around        
        self.summary_ws.cell(row = self._current_line_summary, column = summary_col_editcomment, value=header_summary_edit_comment)
        self.summary_ws.cell(row = self._current_line_summary, column = summary_col_editcomment).border = defs.xl_border_around
        self._current_line_summary += 1
        for i in range(self.count_unit_op):
            self.summary_ws.cell(row=self._current_line_summary, column=summary_col_seq, value=i+1)
            self.summary_ws.cell(row=self._current_line_summary, column=summary_col_seq).border = defs.xl_border_around
            self.summary_ws.cell(row = self._current_line_summary, column = summary_col_uo).border = defs.xl_border_around
            dv_unitops.add(self.summary_ws.cell(row = self._current_line_summary, column = summary_col_uo))
            self.summary_ws.cell(row = self._current_line_summary, column = summary_col_num_subitems).border = defs.xl_border_around
            self.summary_ws.cell(row = self._current_line_summary, column = summary_col_editcomment).border = defs.xl_border_around
            self._current_line_summary += 1
        

    def load_process_summary(self) -> pd.DataFrame:
        df = pd.read_excel(io = self.file_path, sheet_name=sheet_summary_input, header=0)
        self.df_summary = df
        return df

    
    def put_detail_input_form(self, seq: int,specif_header: List[str], menu_dict: Dict[str, List[str]]):
        #This function makes the detail input form based on the number of the unit operations in the process.
        #Prepare options for drop down list(s) called data validation.

        #summary_ws is no loner necessay, as this instance is needed to edit the worksheet before put out to the excel workbook. 
        self.summary_ws = None

        if self.df_summary == None:
            self.load_process_summary()
        num_sub_items = self.df_summary[self.df_summary[header_summary_sequence]==seq][header_summary_num_subitems].item()
        if math.isnan(num_sub_items):
            num_sub_items = 1
        else:
            num_sub_items = int(num_sub_items)
        menu_dict_local: Dict[str, DataValidation] = {}
        for key in menu_dict:
            options = '"'
            for item in menu_dict[key]:
                options += (item+',')
            options = options[:-1]
            options += '"'
            dv_unitops = DataValidation(
                type='list',
                formula1=options,
                allow_blank=True
            )
            menu_dict_local[key]=dv_unitops
            self.detail_ws.add_data_validation(dv_unitops)

        #Note: header[0] == None to align with Excel
        header = [None]+common_header_detail+specif_header
        for col in range(1, len(header), 1):
            self.detail_ws.cell(row = self._current_line_detail, column = col, value=header[col])
            self.detail_ws.cell(row = self._current_line_detail, column = col).border = defs.xl_border_around
            self.detail_ws.cell(row = self._current_line_detail, column = col).font = defs.xl_font_bold
        self._current_line_detail += 1
        
        for count in range(num_sub_items):
            #Note: header[0] == None to align with Excel
            for col in range(1, len(header), 1):
                self.detail_ws.cell(row = self._current_line_detail, column = col).border = defs.xl_border_around
                if header[col] == header_detail_seq:
                    self.detail_ws.cell(row = self._current_line_detail, column = col).value = seq
                if header[col] == header_detail_edit_comment and count == 0:
                    self.detail_ws.cell(row = self._current_line_detail, column = col).value = self.df_summary[self.df_summary[header_summary_sequence]==seq][header_summary_edit_comment].item()
                if (header[col] == header_detail_precomment or header[col] == header_detail_postcomment) and count > 0:
                    self.detail_ws.cell(row = self._current_line_detail, column = col).value = no_comment_instr
                if header[col] in menu_dict_local:
                    temp_key = header[col]
                    menu_dict_local[temp_key].add(self.detail_ws.cell(row = self._current_line_detail, column = col))
            self._current_line_detail +=1
        
        self._current_line_detail +=1

    def load_process_details(self) -> List[pd.DataFrame]:
        crude_df = pd.read_excel(io=self.file_path, sheet_name=sheet_detail_input, header=None)
        temp_list_series: List[pd.Series] = []
        tables: List[pd.DataFrame] = []

        for _, row in crude_df.iterrows():
            if row.isna().all():
                if temp_list_series:
                    tables.append(pd.DataFrame(temp_list_series))
                    temp_list_series.clear()
            else:
                temp_list_series.append(row)
        if temp_list_series:
            tables.append(pd.DataFrame(temp_list_series))
            temp_list_series.clear()

        for i in range(len(tables)):
            tables[i].dropna(axis='columns', how='all', inplace=True)
            tables[i].columns = tables[i].iloc[0]
            tables[i] = (tables[i])[1:].reset_index(drop=True)

        return tables




        
        





                



        