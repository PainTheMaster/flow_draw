import os
import math
import openpyxl as xl
import pandas as pd
import flow_draw.definitions as defs
import flow_draw.materials.materials as mats
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation
from typing import List, Dict

inputfile_base_name = defs.inputfile_base_name
"""Base name for process input Excel file"""

suffix_summary_input_ws = defs.suffix_summary_input_ws
"""suffix for summary input worksheet (tab)"""

suffix_mats_input_ws = defs.mats_suffix_ws

suffix_detail_input_ws = defs.suffix_detail_input_ws
"""suffix for detail input worksheet (tab)"""

header_summary_sequence = defs.header_summary_sequence
header_summary_uo = defs.header_summary_uo
header_summary_num_subitems = defs.header_summary_num_subitems
header_summary_edit_comment = defs.header_summary_edit_comment

summary_col_seq = defs.summary_col_seq
summary_col_uo = defs.summary_col_uo
summary_col_num_subitems = defs.summary_col_num_subitems
summary_col_editcomment = defs.summary_col_editcomment

header_detail_seq = defs.header_detail_seq
header_detail_uo = defs.header_detail_uo
header_detail_edit_comment = header_summary_edit_comment
header_detail_precomment = defs.header_detail_precomment
header_detail_postcomment = defs.header_detail_postcomment

common_header_detail = defs.common_header_detail

no_comment_instr = defs.no_comment_instr

class ProcessIO:
    """
    This class manages IO for process information. Each instance of this class corresponds one-on-one with a process.
    This class has methods to create an input form for summary and details each, load data from the forms, and save the forms as worksheets in an Excel workbook.
    
    Attributes
    -------------
    process_name : str
        The name of the process.
    count_unit_op : int
        The number of unit operations that constitute the process. This must be given from outside the class at first in order to generate the form with proper number of unit operations.
    file_path : str
        The path to the data input form (Excel). The name is automatically generated from the process name. 
    summary_ws : openpyxl.worksheet.worksheet.Worksheet
        Worksheet for summary input.
    detail_ws : openpyxl.worksheet.worksheet.Worksheet
        Worksheet for detail input.
    df_summary : Pandas.DataFrame
        A narrow table for summary input. Holds only sequence numbers, unit operations, and edit comments. 
    _current_line_summary : int
        Current reading/writing line for the summary worksheet. For internal use only.
    _current_line_detail : int
        Current reading/writing line for the detail worksheet. Completely internal.


    """
    def __init__(self, project_name: str, process_name: str, num_unit_op: int):
        """
        Sets the necessary parameters as follows:\n
            -project name
            -process name
            -unit operation number
            -file path to the input file
            -worksheet titles for summary and detail in the file
            -current reading line of summary and detail
        The file name is automatically set as &lt;project name&gt;+&lt;base name="_process_input"&gt;+".xlsx".
        In a similar manner, the worksheet are named after process name (&lt;process name&gt;+&lt;suffix&gt;).
        Please note that only one file is created to gather process and raw materials information for one project. The sole file has sets of worksheets (materials + process summary + process details) to cover all processes. 
        self.__manage_io() is called to newly create the file and worksheets if they don't exist. Otherwise, existing data is loaded.

        Parameters
        ------------
        project_name: str
            The project name.
        
        process_name: str
            The process name.
        
        num_unit_op: int
            The number of the unit operations constituting the process.
        """
        self.project_name: str = project_name
        self.process_name: str = process_name
        self.num_unit_op: int = num_unit_op
        self.file_path = project_name+inputfile_base_name+'.xlsx'
        self.title_summary_ws: str = process_name+suffix_summary_input_ws
        self.summary_ws: Worksheet = None
        self.title_mats_ws: str = process_name+suffix_mats_input_ws
        self.mats_ws: Worksheet = None
        self.title_detail_ws: str = project_name+suffix_detail_input_ws
        self.detail_ws: Worksheet = None
        self.__manage_io()
        self.df_summary: pd.DataFrame = None
        self._current_line_summary:int = 1
        self._current_line_mats:int = 1
        self._current_line_detail:int = 1
    
    def __manage_io(self):
        """
        Creates an Open PyXL workbook for data input, and creates brank work sheets for summary and detail input, if the data input file with the intended name doesn't eixists.
        Otherwise, the files is loaded to re-generate an Open PyXL workbook. Worksheets are created as necessary.
        Please note this only takes care of worksheets, and doesn't handle DataFrame. It's up to other methods such as load_process_summary() or load_process_details().
        """
        if not os.path.isfile(self.file_path):
            self.wb = xl.Workbook()
            self.wb.remove(worksheet=self.wb['Sheet'])
            self.summary_ws: Worksheet = self.wb.create_sheet(title=self.title_summary_ws)
            self.mats_ws: Worksheet = self.wb.create_sheet(title=self.title_mats_ws)
            self.detail_ws: Worksheet = self.wb.create_sheet(title=self.title_detail_ws)
        else:
            self.wb=xl.load_workbook(self.file_path)
            sheet_names = self.wb.sheetnames
            if not self.title_summary_ws in sheet_names:
                self.summary_ws:Worksheet = self.wb.create_sheet(title=self.title_summary_ws)
            else:
                self.summary_ws: Worksheet = self.wb[self.title_summary_ws]
            if not self.title_mats_ws in sheet_names:
                self.mats_ws: Worksheet = self.wb.create_sheet(title=self.title_mats_ws)
            else:
                self.mats_ws: Worksheet =self.wb[self.title_mats_ws]
            if not self.title_detail_ws in sheet_names:
                self.detail_ws: Worksheet = self.wb.create_sheet(title=self.title_detail_ws)
            else:
                self.detail_ws: Worksheet = self.wb[self.title_detail_ws]



    def save_form(self):
        """
        Saves the workbook which holds the worksheets for both summary and details. The output is an Excel file on the storage. No arguments.
        
        Parameters
        -----------
        None

        Returns
        ----------
        None
        """
        self.wb.save(filename=self.file_path)

    def get_path_to_forms(self)->str:
        return self.file_path

    def generate_proc_summary_form(self, list_unit_ops: List[str]):
        """
        Creates a summary input form based on the number of the unit operations in the process.\n
        There will be four coloumns in the newly created form:\n
            -Sequnece Number
            -Unit Operation
            -Nuber of Sub-items
            -Edit Comment
        A pull-down menu will be set in each cell in the Unit Operation column.
        The method only edits the summary_ws in the class, and does not save it as an file on the storage. It's up to another method.

        Parameters
        ---------------
        list_unit_ops: List[str]
            A list of strings which holds the name tags for unit operations, e.g., line_clearance, N2_replacement_ charging, etc. The name tags must be recognizable by the unit_operations.UnitOperation class.
        
        Returns
        ---------------
        None

        Notes
        ---------------
        Requires self.summary_ws: openpyxl.worksheet.worksheet.Worksheet. The worksheet must be available before this method is called. 

        """
        options_dv: str = '"'
        for item in list_unit_ops:
            options_dv += (item+',')
        options_dv = options_dv[:-1]
        options_dv = options_dv+'"'
        dv_unitops = DataValidation(
            type='list',
            formula1=options_dv,
            allow_blank=False
        )
        self.summary_ws.add_data_validation(dv_unitops)

        self._current_line_summary = 1
        self.summary_ws.cell(row = self._current_line_summary, column = summary_col_seq, value=header_summary_sequence)
        self.summary_ws.cell(row = self._current_line_summary, column = summary_col_seq).border = defs.xl_border_around
        self.summary_ws.cell(row = self._current_line_summary, column = summary_col_uo, value=header_summary_uo)
        self.summary_ws.cell(row = self._current_line_summary, column = summary_col_uo).border = defs.xl_border_around
        self.summary_ws.cell(row = self._current_line_summary, column = summary_col_num_subitems, value=header_summary_num_subitems)
        self.summary_ws.cell(row = self._current_line_summary, column = summary_col_num_subitems).border = defs.xl_border_around        
        self.summary_ws.cell(row = self._current_line_summary, column = summary_col_editcomment, value=header_summary_edit_comment)
        self.summary_ws.cell(row = self._current_line_summary, column = summary_col_editcomment).border = defs.xl_border_around
        self._current_line_summary += 1
        for i in range(self.num_unit_op):
            self.summary_ws.cell(row=self._current_line_summary, column=summary_col_seq, value=i+1)
            self.summary_ws.cell(row=self._current_line_summary, column=summary_col_seq).border = defs.xl_border_around
            self.summary_ws.cell(row = self._current_line_summary, column = summary_col_uo).border = defs.xl_border_around
            dv_unitops.add(self.summary_ws.cell(row = self._current_line_summary, column = summary_col_uo))
            self.summary_ws.cell(row = self._current_line_summary, column = summary_col_num_subitems).border = defs.xl_border_around
            self.summary_ws.cell(row = self._current_line_summary, column = summary_col_editcomment).border = defs.xl_border_around
            self._current_line_summary += 1
        

        
    def generate_mats_form(self):
        """
        Generates the input form worksheet for the raw materials for the proces. The worksheet is a part of the input file.

        Parameters
        ------------
        None

        Returns
        ------------
        None
        """
        options_dv: str = f'\"{defs.mats_compo_desig_star},,\"'
        dv_main = DataValidation(
            type='list',
            formula1=options_dv,
            allow_blank=True
        )
        self.mats_ws.add_data_validation(dv_main)
        self._current_line_mats = 1
        self.mats_ws.cell(row=self._current_line_mats, column=defs.mats_col_material).value=defs.mats_header_material
        self.mats_ws.cell(row=self._current_line_mats, column=defs.mats_col_material).border = defs.xl_border_around
        self.mats_ws.cell(row=self._current_line_mats, column=defs.mats_col_main).value=defs.mats_header_main
        self.mats_ws.cell(row=self._current_line_mats, column=defs.mats_col_main).border = defs.xl_border_around
        self.mats_ws.cell(row=self._current_line_mats, column=defs.mats_col_MW).value=defs.mats_header_MW
        self.mats_ws.cell(row=self._current_line_mats, column=defs.mats_col_MW).border = defs.xl_border_around
        self.mats_ws.cell(row=self._current_line_mats, column=defs.mats_col_density).value=defs.mats_header_density
        self.mats_ws.cell(row=self._current_line_mats, column=defs.mats_col_density).border = defs.xl_border_around
        self.mats_ws.cell(row=self._current_line_mats, column=defs.mats_col_conc_assay).value=defs.mats_header_conc_assay
        self.mats_ws.cell(row=self._current_line_mats, column=defs.mats_col_conc_assay).border = defs.xl_border_around
        self.mats_ws.cell(row=self._current_line_mats, column=defs.mats_col_kg_main).value=defs.mats_header_kg_main
        self.mats_ws.cell(row=self._current_line_mats, column=defs.mats_col_kg_main).border = defs.xl_border_around
        self.mats_ws.cell(row=self._current_line_mats, column=defs.mats_col_remark).value=defs.mats_header_remark
        self.mats_ws.cell(row=self._current_line_mats, column=defs.mats_col_remark).border = defs.xl_border_around
        self._current_line_mats += 1
        for _ in range(defs.mats_default_num_rows):
            self.mats_ws.cell(row=self._current_line_mats, column=defs.mats_col_material).border = defs.xl_border_around
            self.mats_ws.cell(row=self._current_line_mats, column=defs.mats_col_main).border = defs.xl_border_around
            dv_main.add(self.mats_ws.cell(row=self._current_line_mats, column=defs.mats_col_main))
            self.mats_ws.cell(row=self._current_line_mats, column=defs.mats_col_MW).border = defs.xl_border_around
            self.mats_ws.cell(row=self._current_line_mats, column=defs.mats_col_density).border = defs.xl_border_around
            self.mats_ws.cell(row=self._current_line_mats, column=defs.mats_col_conc_assay).border = defs.xl_border_around
            self.mats_ws.cell(row=self._current_line_mats, column=defs.mats_col_kg_main).border = defs.xl_border_around
            self.mats_ws.cell(row=self._current_line_mats, column=defs.mats_col_remark).border = defs.xl_border_around
            self._current_line_mats += 1


    def load_process_summary(self) -> pd.DataFrame:
        """
        Loads process summary data (seq, unit operation name, number of subitems, and edit comment) from the summary worksheet in the input Excel file.
        The data is acquired in DataFrame format. The obtained DataFrame object is both stored as an instance variable and returned to the caller.
        The number of the unit operations is counted. process_io.num_unit_op is updated with it so that the generator of the detail input table can make the right nuber of input tables.

        Parameters
        --------------
        None

        Returns
        --------------
        df: pandas.DataFrame
            The summary input data. This is expected to be consists of four columns with a header--Sequence, unit operation name, number of subitems, edit comennt.
        """
        df = pd.read_excel(io = self.file_path, sheet_name=self.title_summary_ws, header=0)
        #count() counts non-NaN items in the column
        self.num_unit_op = df[header_summary_uo].count()
        self.df_summary = df
        return df

    def load_mats(self) -> mats.Materials:
        df =pd.read_excel(io=self.file_path, sheet_name=self.title_mats_ws, header=0)
        mats_this_proc = mats.Materials(df)
        return mats_this_proc
    
    def put_detail_input_table(self, seq: int, specif_header: list[str], menu_dict: dict[str, list[str]]):
        """
        Makes a detail input table for one unit operation.
        Prepares options for drop-down list(s) called "data validation".
        if self.df_summary, from which the unit operation corresponding to the given seq is retrieved, is empty, the method loads the DataFrame object by using self.load_process_summary().
        
        Parameters
        -------------
        seq: int
            Sequnece number of the unit operation.
        
        specif_header: list[str]
            List of header items specific to the unit operation.
        
        menu_dict: dict[str, list[str]]
            (Optional) drop-down items for columns for the unit operation. Each string key is the header item which needs a drop-down list. The list[str] value is the items for the drop-down list.
        """
        #summary_ws is no loner necessay, as this instance is needed to edit the worksheet before put out to the excel workbook. 
        #self.summary_ws does not have to be cleared. As long as it is edited, the contents on the sheet stays intact. The workbook object is responsible for retaining the original data of the ws.
        #self.summary_ws = None

        if self.df_summary == None:
            self.load_process_summary()
        num_sub_items = self.df_summary[self.df_summary[header_summary_sequence]==seq][header_summary_num_subitems].item()

        #perhaps users will omit specifying sub-item numbers if it is 1.
        if math.isnan(num_sub_items):
            num_sub_items = 1
        else:
            num_sub_items = int(num_sub_items)
        menu_dict_local: dict[str, DataValidation] = {}
        for key in menu_dict:
            options = '"'
            #note that menu_dict[key] is a list[str]
            for item in menu_dict[key]:
                options += (item+',')
            #A "," at the end is not needed.
            options = options[:-1]
            options += '"'
            dv_unitops = DataValidation(
                type='list',
                formula1=options,
                allow_blank=True
            )
            menu_dict_local[key]=dv_unitops
            self.detail_ws.add_data_validation(dv_unitops)

        #Note: header[0] == None to align with Excel
        header = [None]+common_header_detail+specif_header
        for col in range(1, len(header), 1):
            self.detail_ws.cell(row = self._current_line_detail, column = col, value=header[col])
            self.detail_ws.cell(row = self._current_line_detail, column = col).border = defs.xl_border_around
            self.detail_ws.cell(row = self._current_line_detail, column = col).font = defs.xl_font_bold
        self._current_line_detail += 1
        
        for count in range(num_sub_items):
            #Note: header[0] == None in order to align with Excel
            for col in range(1, len(header), 1):
                self.detail_ws.cell(row = self._current_line_detail, column = col).border = defs.xl_border_around
                if header[col] == header_detail_seq:
                    self.detail_ws.cell(row = self._current_line_detail, column = col).value = seq
                if header[col] == header_detail_uo:
                    self.detail_ws.cell(row = self._current_line_detail, column = col).value = self.df_summary[self.df_summary[header_summary_sequence]==seq][header_summary_uo].item()
                if header[col] == header_detail_edit_comment and count == 0:
                    self.detail_ws.cell(row = self._current_line_detail, column = col).value = self.df_summary[self.df_summary[header_summary_sequence]==seq][header_summary_edit_comment].item()
                if (header[col] == header_detail_precomment or header[col] == header_detail_postcomment) and count > 0:
                    self.detail_ws.cell(row = self._current_line_detail, column = col).value = no_comment_instr
                if header[col] in menu_dict_local:
                    temp_key = header[col]
                    menu_dict_local[temp_key].add(self.detail_ws.cell(row = self._current_line_detail, column = col))
            self._current_line_detail +=1

        self._current_line_detail +=1

    def load_process_details(self) -> list[pd.DataFrame]:
        """
        Processes the detail input sheet in an Excel file into a collection of DataFrame objects, each of which represents a single unit operation.
        The input Excel sheet contains multiple tables with various width. This method decomposes it more UnitOperation-object-friendly form.

        Parameters
        ------------
            None

        Returns
        ------------
        tables: list[pd.DataFrame]
            Each value of which is a DataFrame object for a single unit operation in the process.
        """
        crude_df = pd.read_excel(io=self.file_path, sheet_name=self.title_detail_ws, header=None)
        temp_list_series: list[pd.Series] = [] #Collection of lines (rows) cut out from a valid table. Temporal.
        tables: list[pd.DataFrame] = []

        for _, row in crude_df.iterrows():
            if row.isna().all():
                if temp_list_series:
                    tables.append(pd.DataFrame(temp_list_series))
                    temp_list_series.clear()
            else:
                temp_list_series.append(row)
        if temp_list_series:
            tables.append(pd.DataFrame(temp_list_series))
            temp_list_series.clear()

        for i in range(len(tables)):
            tables[i].dropna(axis='columns', how='all', inplace=True)
            tables[i].columns = tables[i].iloc[0]
            tables[i] = (tables[i])[1:].reset_index(drop=True)

        return tables




        
        





                



        