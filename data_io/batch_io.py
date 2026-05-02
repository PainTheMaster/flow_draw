import os
import openpyxl as xl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
import flow_draw.definitions as defs



outline_file_path:str = defs.src_io_batch_input_file_name
outline_ws_name:str = defs.src_io_batch_outline_ws
col_label_item:str = defs.col_io_batch_item_ouln_tab
col_label_value:str = defs.col_io_batch_value_ouln_tab

row_hedr:int = defs.row_io_batch_hedr_ouln_tab
col_item:int = defs.col_io_batch_item_ouln_tab
col_val:int = defs.col_io_batch_value_ouln_tab

row_hedr_excel:int = row_hedr+1   #The meaning is in parallel with DataFrame, but the counting start with 1 in Excel.
col_item_excel:int = col_item+1   #The meaning is in parallel with DataFrame, but the counting start with 1 in Excel.
col_val_excel:int = col_val+1    #The meaning is in parallel with DataFrame, but the counting start with 1 in Excel.

hedr_batch_item:str = defs.hedr_io_batch_item
hedr_batch_value:str = defs.hedr_io_batch_value

item_batch_name:str = defs.item_io_batch_batch_name
item_batch_remark:str = defs.item_io_batch_batch_remark
item_proc_name_stem:str = defs.item_io_batch_proc_name_stem
item_proc_count_uo_stem:str = defs.item_io_batch_proc_count_uo_stem
item_proc_remark_stem:str = defs.item_io_batch_proc_remark_stem

dflt_num_procs:int = defs.dflt_io_batch_num_procs

class BatchIO:
    def __init__(self):
        self.outline_file_path:str = outline_file_path
        self.title_outline_ws:str = outline_ws_name 
        self.wb:Workbook = None
        self.ws:Worksheet = None
        self.num_procs:int = None
        self.current_line_ws:int = 1
    
    def load_outline(self) -> pd.DataFrame:
        """
        Read from the input file and transform it into a pandas.DataFrame. The header item is 
        """
        #col_labels: list[int] = [col_label_item, col_label_value]
        df_batch = pd.read_excel(io=self.outline_file_path, sheet_name=self.title_outline_ws, header=row_hedr, index_col=col_item)
        return df_batch
        
    def __manage_io(self):
        """Purpose: get the workbook and worksheet ready. At the end of the method, a proper workbook and worksheet are registered as instance variables"""
        if not os.path.isfile(path=self.outline_file_path):
            self.wb = xl.Workbook()
            self.wb.remove(worksheet=self.wb["Sheet"])
            self.ws = self.wb.create_sheet(title=self.title_outline_ws)
        else:
            self.wb = xl.load_workbook(filename=self.outline_file_path)
            if not self.title_outline_ws in self.wb.worksheets:
                self.ws = self.wb.create_sheet(title=self.title_outline_ws)
            else:
                self.ws = self.wb[self.title_outline_ws]
    
    def save_wb(self):
        self.wb.save(self.outline_file_path)

    def generate_form(self):
        self.__manage_io()
        if self.num_procs is None:
            self.num_procs = dflt_num_procs
        #Header
        self.current_line_ws = row_hedr_excel
        self.ws.cell(row=self.current_line_ws, column=col_item_excel, value=hedr_batch_item)
        self.ws.cell(row=self.current_line_ws, column=col_item_excel).border = defs.xl_border_around
        self.ws.cell(row=self.current_line_ws, column=col_val_excel, value=hedr_batch_value)
        self.ws.cell(row=self.current_line_ws, column=col_val_excel).border = defs.xl_border_around
        self.current_line_ws += 1

        #Batch outline
        self.ws.cell(row=self.current_line_ws, column=col_item_excel, value=item_batch_name)
        self.ws.cell(row=self.current_line_ws, column=col_item_excel).border = defs.xl_border_around
        self.ws.cell(row=self.current_line_ws, column=col_val_excel).border = defs.xl_border_around
        self.current_line_ws += 1
        self.ws.cell(row=self.current_line_ws, column=col_item_excel, value=item_batch_remark)
        self.ws.cell(row=self.current_line_ws, column=col_item_excel).border = defs.xl_border_around
        self.ws.cell(row=self.current_line_ws, column=col_val_excel).border = defs.xl_border_around
        self.current_line_ws += 1

        for i in range(1, self.num_procs+1):
            item_proc_name = item_proc_name_stem.format(i)
            item_uo_count = item_proc_count_uo_stem.format(i)
            item_proc_remark = item_proc_remark_stem.format(i)
            self.ws.cell(row=self.current_line_ws, column=col_item_excel, value=item_proc_name)
            self.ws.cell(row=self.current_line_ws, column=col_item_excel).border = defs.xl_border_around
            self.ws.cell(row=self.current_line_ws, column=col_val_excel).border = defs.xl_border_around
            self.current_line_ws += 1
            self.ws.cell(row=self.current_line_ws, column=col_item_excel, value=item_uo_count)
            self.ws.cell(row=self.current_line_ws, column=col_item_excel).border = defs.xl_border_around
            self.ws.cell(row=self.current_line_ws, column=col_val_excel).border = defs.xl_border_around
            self.current_line_ws += 1
            self.ws.cell(row=self.current_line_ws, column=col_item_excel, value=item_proc_remark)
            self.ws.cell(row=self.current_line_ws, column=col_item_excel).border = defs.xl_border_around
            self.ws.cell(row=self.current_line_ws, column=col_val_excel).border = defs.xl_border_around
            self.current_line_ws += 1
