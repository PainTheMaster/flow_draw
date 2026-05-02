import unittest
import os
import openpyxl as xl
from openpyxl.worksheet.worksheet import Worksheet
import flow_draw.definitions as defs
import flow_draw.batch.batch as batch
import flow_draw.data_io.batch_io as batch_io
import flow_draw.batch.process.process as proc
import flow_draw.data_io.process_io as proc_io
import flow_draw.batch.process.unit_operations as uos
import flow_draw.batch.process.unit_operations.charging as charging



class TestIO(unittest.TestCase):
    def setUp(self):
        self.batch_name = "test_batch_0000"
        self.process_name = "test_process"
        self.test_proc_io = proc_io.ProcessIO(batch_name=self.batch_name, process_name=self.process_name, num_unit_op=1)
        self.test_proc_io.generate_proc_summary_form(uos.unit_operation.list_unit_ops)
        self.test_proc_io.generate_mats_form()
        self.test_proc_io.save_form()    
        return super().setUp()
    

    def mats_test_data_row_injection(self, row:int = 2,
                                    name_mat: str ="",
                                    is_main: bool=False,
                                    mw: float=None,
                                    density: float=None,
                                    assay_pct: float=None,
                                    kg: float=None,
                                    remark: str=""):
        """
        Helper method. Injects specific values to the input form (an worksheet in an Excel file on the hard drive)
        Don't forget to save the file!
        """
        ws_chem = self.test_proc_io.mats_ws
        ws_chem.cell(row=row, column=defs.col_nr_io_mats_mat).value=name_mat
        if is_main:
            ws_chem.cell(row=row, column=defs.col_nr_io_mats_main).value=defs.itm_io_mats_desig_star
        if mw:
            ws_chem.cell(row=row, column=defs.col_nr_io_mats_mw).value=mw 
        if density:
            ws_chem.cell(row=row, column=defs.col_nr_io_mats_dnsty).value=density
        if assay_pct:
            ws_chem.cell(row=row, column=defs.col_nr_io_mats_concasy).value=assay_pct
        if kg:
            ws_chem.cell(row=row, column=defs.col_nr_io_mats_kgmain).value=kg
        ws_chem.cell(row=row, column=defs.col_nr_io_mats_remark).value=remark
        #self.test_proc_io.save_form()
    
    def summary_test_data_row_injection(self, row:int = 2,
                                        seq: int =None,
                                        uo: str="",
                                        num_subitem: int=None,
                                        editcomment: str=""):
        """
        Helper method. Injects specific values to the input form (an worksheet in an Excel file on the hard drive)
        Don't forget to save the file!
        """
        ws_summary = self.test_proc_io.summary_ws
        if seq:
            ws_summary.cell(row=row, column=defs.col_nr_io_sumry_seq).value=seq
        if uo:
            ws_summary.cell(row=row, column=defs.col_nr_io_sumry_uo).value=uo
        if num_subitem:
            ws_summary.cell(row=row, column=defs.col_nr_io_sumry_num_subitms).value=num_subitem
        ws_summary.cell(row=row, column=defs.col_nr_io_sumry_edt_cmnt).value=editcomment


    def test_0001_input_generation(self):
        """Checks if the right input excel form is generated. The form generators for summary d materials are called in self.setUp()"""
        print("expected filename: ", self.batch_name+defs.src_io_filebasename+".xlsx")
        self.assertTrue(os.path.isfile(self.batch_name+defs.src_io_filebasename+".xlsx"))
    
    def test_0002_uo_deriv_name_registr(self):
        """
        Checks if the name of each unit operation class is registerd. The registry is located in the module batch.process.unit_operations.unit_operation.
        The registration takes place when the class (not an instance) is generated.
        """
        self.assertIn(member=defs.tag_uo_charging, container=uos.unit_operation.list_unit_ops)

    def test_0003_uo_deriv_class_registr(self):
        """
        Checks if each sort of unit operation class is registerd. The registry is located in the module batch.process.unit_operations.unit_operation.
        The registration takes place when the class (not an instance) is generated.
        """
        self.assertIn(member=charging.Charging, container=uos.unit_operation.registry_uo_cls.values())

    def test_0004_materials_form(self):
        self.mats_test_data_row_injection(row=2,
                                          name_mat="NaCl",
                                          is_main=True,
                                          mw=58.44,
                                          density=2.17,
                                          assay_pct=100.0,
                                          kg=1.0,
                                          remark="Table salt.")
        self.mats_test_data_row_injection(row=3,
                                          name_mat="Water",
                                          is_main=False,
                                          mw=18.01,
                                          density=1.00,
                                          assay_pct=100.0,
                                          remark="Pure water.")
        self.test_proc_io.save_form()


class TestForProcessCls(unittest.TestCase):
    def setUp(self):
        self.batch_name = "test_batch_1000"
        self.process_name = "test_process"
        self.num_uo = 2
        self.test_process = proc.Process(batch_name=self.batch_name, process_name=self.process_name, num_uo=self.num_uo, comment="dummy comment")
        self.test_process.put_summary_mats_input_form()


        # self.test_proc_io = proc_io.ProcessIO(batch_name=self.batch_name, process_name=self.process_name, num_unit_op=1)
        # self.test_proc_io.generate_proc_summary_form(uos.unit_operation.list_unit_ops)
        # self.test_proc_io.generate_mats_form()
        # self.test_proc_io.save_form()    
        # return super().setUp()        


    def mats_test_data_row_injection(self, row:int = 2,
                                    name_mat: str ="",
                                    is_main: bool=False,
                                    mw: float=None,
                                    density: float=None,
                                    assay_pct: float=None,
                                    kg: float=None,
                                    remark: str=""):
        """
        Helper method. Injects specific values to the input form (an worksheet in an Excel file on the hard drive)
        Don't forget to save the file!
        """
        data_input = self.test_process.data_input
        ws_chem = data_input.mats_ws
        ws_chem.cell(row=row, column=defs.col_nr_io_mats_mat).value=name_mat
        if is_main:
            ws_chem.cell(row=row, column=defs.col_nr_io_mats_main).value=defs.itm_io_mats_desig_star
        if mw:
            ws_chem.cell(row=row, column=defs.col_nr_io_mats_mw).value=mw 
        if density:
            ws_chem.cell(row=row, column=defs.col_nr_io_mats_dnsty).value=density
        if assay_pct:
            ws_chem.cell(row=row, column=defs.col_nr_io_mats_concasy).value=assay_pct
        if kg:
            ws_chem.cell(row=row, column=defs.col_nr_io_mats_kgmain).value=kg
        ws_chem.cell(row=row, column=defs.col_nr_io_mats_remark).value=remark
        #self.test_proc_io.save_form()
    
    def summary_test_data_row_injection(self, row:int = 2,
                                        seq: int =None,
                                        uo: str="",
                                        num_subitem: int=None,
                                        editcomment: str=""):
        """
        Helper method. Injects specific values to the input form (an worksheet in an Excel file on the hard drive)
        Don't forget to save the file!
        """
        data_input = self.test_process.data_input
        ws_summary = data_input.summary_ws
        if seq:
            ws_summary.cell(row=row, column=defs.col_nr_io_sumry_seq).value=seq
        if uo:
            ws_summary.cell(row=row, column=defs.col_nr_io_sumry_uo).value=uo
        if num_subitem:
            ws_summary.cell(row=row, column=defs.col_nr_io_sumry_num_subitms).value=num_subitem
        ws_summary.cell(row=row, column=defs.col_nr_io_sumry_edt_cmnt).value=editcomment

    def save_injected_data(self):
        data_input = self.test_process.data_input
        data_input.save_form()


    def test_1000_proc_generated(self):
        tp = self.test_process
        check_result = ((tp.batch_name == self.batch_name) and
                        (tp.process_name == self.process_name) and
                        (tp.num_uo == self.num_uo))
        self.assertTrue(check_result)
    
    def test_1001_summary_write_read_consistency(self):
        tp=self.test_process

        seq_proc_1 = 1
        uo_proc_1 = defs.tag_uo_charging
        num_subitem_proc_1 = 2
        editcomment_proc_1 = "test process with two sub items 1"
        self.summary_test_data_row_injection(row=2,
                                             seq=seq_proc_1,
                                             uo=uo_proc_1,
                                             num_subitem=num_subitem_proc_1,
                                             editcomment=editcomment_proc_1)

        seq_proc_2 = 2
        uo_proc_2 = defs.tag_uo_charging
        num_subitem_proc_2 = 1
        editcomment_proc_2 = "test process with only one sub item"
        self.summary_test_data_row_injection(row=3,
                                             seq=seq_proc_2,
                                             uo=uo_proc_2,
                                             num_subitem=num_subitem_proc_2,
                                             editcomment=editcomment_proc_2)

        name_NaCl="NaCl"
        main_NaCl=True
        mw_NaCl=58.44
        density_NaCl=2.17
        assay_pct_NaCl=100.0
        kg_NaCl=1.0
        remark_NaCl="Table salt."
        self.mats_test_data_row_injection(row=2,
                                          name_mat=name_NaCl,
                                          is_main=main_NaCl,
                                          mw=mw_NaCl,
                                          density=density_NaCl,
                                          assay_pct=assay_pct_NaCl,
                                          kg=kg_NaCl,
                                          remark=remark_NaCl)
        
        name_water="Water"
        main_water=False
        mw_Water=18.01
        density_water=1.00
        assay_pct_water=100.0
        remark_water="Pure water."
        self.mats_test_data_row_injection(row=3,
                                          name_mat=name_water,
                                          is_main=main_water,
                                          mw=mw_Water,
                                          density=density_water,
                                          assay_pct=assay_pct_water,
                                          remark=remark_water)
        
        self.save_injected_data()

        tp.load_uo_summary()
        tp.load_materials_data()


        result_num_uos = (tp.num_uo == 2)

        uo_loaded_1 = tp.list_uo[0]
        result_proc_1 = ((uo_loaded_1.num_subitems == num_subitem_proc_1) and
                       (uo_loaded_1.uo_name == uo_proc_1) and
                       (uo_loaded_1.operation_seq==seq_proc_1) and
                       (uo_loaded_1.num_subitems==num_subitem_proc_1)and
                       (uo_loaded_1.edit_comment==editcomment_proc_1))
        
        uo_loaded_2 = tp.list_uo[1]
        result_proc_2 = ((uo_loaded_2.num_subitems == num_subitem_proc_2) and
                       (uo_loaded_2.uo_name == uo_proc_2) and
                       (uo_loaded_2.operation_seq==seq_proc_2) and
                       (uo_loaded_2.num_subitems==num_subitem_proc_2)and
                       (uo_loaded_2.edit_comment==editcomment_proc_2))
        
        mats = tp.mats_data
        temp_mol_NaCl = kg_NaCl*1000.0/mw_NaCl
        result_mats1 = ((mats.kg_main_mat==kg_NaCl) and
                        (temp_mol_NaCl*0.9999<mats.mol_main_mat) and (mats.mol_main_mat < temp_mol_NaCl*1.0001))
        
        self.assertTrue(result_num_uos and result_proc_1 and result_proc_2 and result_mats1)



class TestForBatchCls(unittest.TestCase):
    def setUp(self):
        self.batch_input = batch_io.BatchIO()
        self.batch_input.generate_form()
        self.inject_test_data(self.batch_input)
        self.batch_input.save_wb()
        self.test_batch = batch.Batch()
        return super().setUp()
    
    def test_2000_load_from_inputform(self):
        df_outline = self.batch_input.load_outline()
        self.test_batch.load_outline(df_batch_outline=df_outline)
        proc = self.test_batch.list_proc[0]
        test_result = ((self.test_batch.batch_name == "batch_test_2000") and
                       (self.test_batch.batch_comment == "remark for test 2000") and
                       (self.test_batch.num_procs == 1) and
                       (proc.process_name == "P1_test2000") and
                       (proc.num_uo == 3))
        self.assertTrue(test_result)
    
    def inject_test_data(self, bio: batch_io.BatchIO):
        ws: Worksheet = bio.ws
        ws.cell(row=2, column=2, value="batch_test_2000")
        ws.cell(row=3, column=2, value="remark for test 2000")
        ws.cell(row=4, column=2, value="P1_test2000")
        ws.cell(row=5, column=2, value=3)        
        ws.cell(row=6, column=2, value="Remark P1 test2000")           
        





if __name__=="__main__":
    unittest.main()
